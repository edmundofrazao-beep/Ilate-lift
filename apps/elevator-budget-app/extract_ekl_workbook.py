from __future__ import annotations

import json
from pathlib import Path
import sys
from typing import Any

import pandas as pd
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parent.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from elevator_budget_app.ekl_orc_workbook import (
    CLIENT_SHEET,
    DEFAULT_WORKBOOK_PATH,
    MODEL_SHEET,
    OVERVIEW_SHEET,
)

OUTPUT_DIR = ROOT_DIR / "data" / "extracted"
REPORT_PATH = OUTPUT_DIR / "ekl_extraction_report.md"


def _json_dump(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def _read_cell(ws, cell: str) -> Any:
    return ws[cell].value


def workbook_inventory() -> list[dict[str, Any]]:
    return [
        {"sheet": "01-DB-RSM Means EN-US", "decision": "reference", "reason": "Base RSM em EN-US; útil para rastrear origem, não para consumo direto na app."},
        {"sheet": "02-DB-NEE EN-US", "decision": "reference", "reason": "Base NEE original em imperial/EN-US; manter como referência de proveniência."},
        {"sheet": "02-DB-NEE PT", "decision": "extract", "reason": "Catálogo convertido com produtividade e unidades adaptadas; útil como fonte de linhas tipo."},
        {"sheet": "03-DB-Quantum PT", "decision": "extract", "reason": "Catálogo alargado de produtos e coeficientes laborais; útil como referência de catálogo."},
        {"sheet": "03-DB-Quantum EN-US", "decision": "reference", "reason": "Fonte original em EN-US."},
        {"sheet": "g3-1", "decision": "ignore", "reason": "Folha intermédia/rascunho sem valor claro de integração imediata."},
        {"sheet": "g3-4", "decision": "ignore", "reason": "Folha intermédia/rascunho sem valor claro de integração imediata."},
        {"sheet": "01-DB-RSM Means PTv0.3", "decision": "reference", "reason": "Base PT preliminar, útil como histórico mas não prioritária para integração imediata."},
        {"sheet": "US--PT", "decision": "extract", "reason": "Parâmetros de contextualização de custos US vs PT; útil como metadado metodológico."},
        {"sheet": "OECD.Stat export", "decision": "reference", "reason": "Fonte bruta de índices; guardar só como suporte metodológico."},
        {"sheet": "$-->€", "decision": "extract", "reason": "Conversão cambial histórica reutilizável como referência de metodologia."},
        {"sheet": "04-VPMO->m(kg)", "decision": "reference", "reason": "Conversões técnicas auxiliares; melhor manter como referência do que integrar diretamente."},
        {"sheet": "05-mm2->m(kg)", "decision": "reference", "reason": "Conversões técnicas auxiliares; útil para rastreio mas não crítica no motor atual."},
        {"sheet": "06-MO 2020 (TBD)", "decision": "extract", "reason": "Perfis e tarifas de mão de obra reutilizáveis."},
        {"sheet": "11 - RSM 2020-Crews-metric", "decision": "extract", "reason": "Estruturas de crew e custos/labor-hour úteis como benchmark."},
        {"sheet": "12 - RSM 2020-Crews-Maintenance", "decision": "extract", "reason": "Benchmark de crews de manutenção."},
        {"sheet": "14 - RSM 2020-Equipment rental", "decision": "extract", "reason": "Referência para aluguer de equipamento e logística."},
        {"sheet": "08-Cabos", "decision": "extract", "reason": "Tabela técnica de tempos e secções de cabo, útil para microcatálogo."},
        {"sheet": "2020 ORC MODEL BASE v.06-msi", "decision": "reference", "reason": "Versão base do modelo, mas a folha 2020 - EKL ORC MODEL 0.5 é mais próxima do uso operacional."},
        {"sheet": "10-Manpower Required", "decision": "extract", "reason": "Regras e conversões de homem-hora para cabos e unidades; útil como catálogo técnico auxiliar."},
        {"sheet": "2016 - 2018", "decision": "reference", "reason": "Histórico de preços/coeficientes; guardar apenas para auditoria."},
        {"sheet": "2019-MSI", "decision": "extract", "reason": "Linhas tipo com material, HH e preços agregados; bom candidato a catálogo legado."},
        {"sheet": "2020-MSI", "decision": "extract", "reason": "Versão mais recente de linhas MSI, útil como catálogo legado."},
        {"sheet": "2020-COEM ORC MOD-0.1", "decision": "reference", "reason": "Versão intermédia de modelo orçamental; útil só como histórico."},
        {"sheet": "2020 - EKL ORC MODEL 0.5", "decision": "extract", "reason": "Núcleo do modelo: parâmetros, taxonomia de custos e configuração comercial."},
        {"sheet": "ELECT+INST", "decision": "extract", "reason": "Mapa cliente tipo com estrutura de linhas e unidades."},
        {"sheet": "2020 - EKL COST OVERVIEW 0.3", "decision": "extract", "reason": "Resumo orçamental, custos logísticos e saídas relevantes."},
        {"sheet": "2020 - EKL LISTA IOs - PHC", "decision": "extract", "reason": "Taxonomia de inputs/outputs do modelo."},
        {"sheet": "2020 - Ex ORC Cliente.", "decision": "extract", "reason": "Mapa cliente tipo com linhas de orçamento e quantidades."},
        {"sheet": "AA - Custos Por Viatura", "decision": "extract", "reason": "Referência de custos de viatura/logística."},
    ]


def extract_parametros_comerciais(wb_values, wb_formulas) -> dict[str, Any]:
    model_v = wb_values[MODEL_SHEET]
    model_f = wb_formulas[MODEL_SHEET]
    overview_v = wb_values[OVERVIEW_SHEET]
    values = {
        "fonte": MODEL_SHEET,
        "parametros_modelo": [
            {"key": "margem_material", "label": "Margem material", "sheet": MODEL_SHEET, "cell": "O21", "value": _read_cell(model_v, "O21")},
            {"key": "margem_mao_obra", "label": "Margem mão de obra", "sheet": MODEL_SHEET, "cell": "P21", "value": _read_cell(model_v, "P21")},
            {"key": "desconto_cliente", "label": "Desconto cliente", "sheet": MODEL_SHEET, "cell": "Q21", "value": _read_cell(model_v, "Q21")},
            {"key": "numero_homens", "label": "Número de homens", "sheet": MODEL_SHEET, "cell": "T21", "value": _read_cell(model_v, "T21")},
            {"key": "horas_semanais", "label": "Horas semanais", "sheet": MODEL_SHEET, "cell": "U21", "value": _read_cell(model_v, "U21")},
            {"key": "natureza_servico", "label": "Natureza do serviço", "sheet": MODEL_SHEET, "cell": "O26", "value": _read_cell(model_v, "O26")},
        ],
        "custos_logisticos": [
            {"key": "deslocacoes_minimas", "label": "Mínimo deslocações", "sheet": OVERVIEW_SHEET, "cell": "C38", "value": _read_cell(overview_v, "C38")},
            {"key": "distancia_km", "label": "Distância km", "sheet": OVERVIEW_SHEET, "cell": "D38", "value": _read_cell(overview_v, "D38")},
            {"key": "desgaste_por_km", "label": "Desgaste / km", "sheet": OVERVIEW_SHEET, "cell": "E38", "value": _read_cell(overview_v, "E38")},
            {"key": "combustivel_l_100km", "label": "Combustível / 100 km", "sheet": OVERVIEW_SHEET, "cell": "F38", "value": _read_cell(overview_v, "F38")},
            {"key": "preco_combustivel_l", "label": "Preço combustível / litro", "sheet": OVERVIEW_SHEET, "cell": "G38", "value": _read_cell(overview_v, "G38")},
            {"key": "hotel_preco_noite", "label": "Preço hotel / noite / pessoa", "sheet": OVERVIEW_SHEET, "cell": "D44", "value": _read_cell(overview_v, "D44")},
            {"key": "ajuda_custo_km_homem", "label": "Ajudas de custo / km / homem", "sheet": OVERVIEW_SHEET, "cell": "C50", "value": _read_cell(overview_v, "C50")},
        ],
        "metadados_documento": [
            {"key": "documento_cliente", "sheet": MODEL_SHEET, "cell": "K21", "value": _read_cell(model_v, "K21"), "formula": _read_cell(model_f, "K21")},
            {"key": "nome_empreitada", "sheet": MODEL_SHEET, "cell": "D22", "value": _read_cell(model_v, "D22"), "formula": _read_cell(model_f, "D22")},
            {"key": "tipo_empreitada", "sheet": MODEL_SHEET, "cell": "D23", "value": _read_cell(model_v, "D23"), "formula": _read_cell(model_f, "D23")},
        ],
    }
    return values


def extract_tarifas_mao_obra(wb_values) -> dict[str, Any]:
    model_ws = wb_values[MODEL_SHEET]
    mo_ws = wb_values["06-MO 2020 (TBD)"]
    tariffs = []
    for row_idx in range(8, 17):
        tariffs.append(
            {
                "perfil": _clean_text(model_ws[f"AU{row_idx}"].value),
                "tarifa_modelo_2020_eur_h": model_ws[f"AV{row_idx}"].value,
                "sheet": MODEL_SHEET,
                "cell": f"AV{row_idx}",
            }
        )
    benchmarks = []
    for row_idx in range(6, 13):
        benchmarks.append(
            {
                "codigo": _clean_text(mo_ws[f"C{row_idx}"].value),
                "perfil": _clean_text(mo_ws[f"D{row_idx}"].value),
                "unidade": _clean_text(mo_ws[f"E{row_idx}"].value),
                "referencia_interna": mo_ws[f"F{row_idx}"].value,
                "navigator_pvp_3_eur_h": mo_ws[f"I{row_idx}"].value,
                "mecwide_pvp_5_eur_h": mo_ws[f"K{row_idx}"].value,
                "schneider_pvp_6_eur_h": mo_ws[f"L{row_idx}"].value,
                "galp_pvp_7_eur_h": mo_ws[f"M{row_idx}"].value,
                "europa_pvp_7_eur_h": mo_ws[f"O{row_idx}"].value,
                "us_canada_pvp_8_eur_h": mo_ws[f"P{row_idx}"].value,
                "mocambique_pvp_9_eur_h": mo_ws[f"Q{row_idx}"].value,
                "angola_pvp_10_eur_h": mo_ws[f"R{row_idx}"].value,
            }
        )
    return {"tarifas_modelo": tariffs, "benchmarks_2020": benchmarks}


def extract_categorias_orcamento(wb_values) -> dict[str, Any]:
    model_ws = wb_values[MODEL_SHEET]
    client_ws = wb_values[CLIENT_SHEET]
    cost_taxonomy = []
    category_specs = [
        ("00", "FORNECEDORES"),
        ("01", "MATERIAIS"),
        ("02", "MÃO DE OBRA"),
        ("03", "CUSTOS FINAIS"),
        ("04", "VENDAS FINAIS"),
        ("05", "AJUSTES FINAIS"),
        ("06", "TOTAL HH POR CATEGORIA"),
        ("07", "CUSTOS TOTAIS POR CATEGORIA"),
        ("08", "VENDAS TOTAIS POR CATEGORIA"),
        ("09", "RESUMO FINAL"),
    ]
    for code, label in category_specs:
        cost_taxonomy.append({"codigo": code, "label": label, "source": MODEL_SHEET})

    service_sections = []
    for row_idx in range(20, min(client_ws.max_row, 120) + 1):
        item = client_ws[f"A{row_idx}"].value
        desc = _clean_text(client_ws[f"B{row_idx}"].value)
        unit = _clean_text(client_ws[f"C{row_idx}"].value)
        qty = client_ws[f"D{row_idx}"].value
        if not item or not desc:
            continue
        if qty is None and unit in {"", "TOT.C.", "TOT.T.", "Vg"}:
            service_sections.append(
                {
                    "codigo": str(item),
                    "descricao": desc,
                    "unidade": unit or None,
                    "tipo": "secao" if "." not in str(item) or str(item).endswith(".0") else "grupo",
                }
            )
    return {"taxonomia_custos": cost_taxonomy, "secoes_servico": service_sections}


def extract_linhas_tipo_orcamento(wb_values) -> pd.DataFrame:
    client_ws = wb_values[CLIENT_SHEET]
    rows = []
    current_section = ""
    for row_idx in range(20, min(client_ws.max_row, 140) + 1):
        item = client_ws[f"A{row_idx}"].value
        desc = _clean_text(client_ws[f"B{row_idx}"].value)
        unit = _clean_text(client_ws[f"C{row_idx}"].value)
        qty = client_ws[f"D{row_idx}"].value
        unit_price = client_ws[f"E{row_idx}"].value
        total = client_ws[f"F{row_idx}"].value
        if not item or not desc:
            continue
        if qty is None and unit in {"", "TOT.C.", "TOT.T.", "Vg"}:
            current_section = desc
        rows.append(
            {
                "codigo": str(item),
                "descricao": desc,
                "secao_pai": current_section or None,
                "unidade": unit or None,
                "quantidade_referencia": qty,
                "preco_unitario_cache": unit_price,
                "total_cache": total,
                "tipo_linha": "item" if qty is not None else "cabecalho",
                "sheet": CLIENT_SHEET,
                "row": row_idx,
            }
        )
    return pd.DataFrame(rows)


def extract_resumo_orcamental(wb_values) -> dict[str, Any]:
    model_ws = wb_values[MODEL_SHEET]
    overview_ws = wb_values[OVERVIEW_SHEET]
    client_ws = wb_values[CLIENT_SHEET]
    summary = {
        "documento": {
            "empresa": client_ws["A1"].value,
            "cliente": client_ws["A2"].value,
            "empreitada": client_ws["A3"].value,
            "tipo_empreitada": client_ws["A5"].value,
            "documento_cliente": client_ws["A6"].value,
        },
        "resumo_valores": [
            {"label": "Total cliente", "sheet": CLIENT_SHEET, "cell": "F12", "value": client_ws["F12"].value},
            {"label": "Total com desconto", "sheet": MODEL_SHEET, "cell": "Q22", "value": model_ws["Q22"].value},
            {"label": "Total sem desconto", "sheet": MODEL_SHEET, "cell": "Q23", "value": model_ws["Q23"].value},
            {"label": "Homem-hora total", "sheet": MODEL_SHEET, "cell": "AS21", "value": model_ws["AS21"].value},
            {"label": "Execução prevista (dias)", "sheet": OVERVIEW_SHEET, "cell": "M14", "value": overview_ws["M14"].value},
            {"label": "Custos totais MO", "sheet": OVERVIEW_SHEET, "cell": "M26", "value": overview_ws["M26"].value},
            {"label": "Vendas totais MO", "sheet": OVERVIEW_SHEET, "cell": "M32", "value": overview_ws["M32"].value},
            {"label": "Custos viaturas", "sheet": OVERVIEW_SHEET, "cell": "M38", "value": overview_ws["M38"].value},
            {"label": "Custos pernoita", "sheet": OVERVIEW_SHEET, "cell": "M44", "value": overview_ws["M44"].value},
            {"label": "Ajudas de custo", "sheet": OVERVIEW_SHEET, "cell": "M50", "value": overview_ws["M50"].value},
        ],
    }
    return summary


def extract_catalogos_auxiliares(wb_values) -> dict[str, pd.DataFrame]:
    cable_ws = wb_values["08-Cabos"]
    cable_rows = []
    for row_idx in range(4, min(cable_ws.max_row, 40) + 1):
        if cable_ws[f"B{row_idx}"].value is None:
            continue
        cable_rows.append(
            {
                "n_condutores": cable_ws[f"B{row_idx}"].value,
                "secao_mm2": cable_ws[f"C{row_idx}"].value,
                "diametro_exterior_vv": cable_ws[f"D{row_idx}"].value,
                "diametro_exterior_vmv": cable_ws[f"E{row_idx}"].value,
                "tempo_vala_h": cable_ws[f"F{row_idx}"].value,
                "tempo_calha_h": cable_ws[f"G{row_idx}"].value,
                "tempo_tubo_h": cable_ws[f"H{row_idx}"].value,
                "tempo_abracadeira_h": cable_ws[f"I{row_idx}"].value,
                "tempo_ligacao_uma_ponta_h": cable_ws[f"J{row_idx}"].value,
                "tempo_ligacao_duas_pontas_h": cable_ws[f"K{row_idx}"].value,
                "awg_ref": cable_ws[f"N{row_idx}"].value,
            }
        )
    manpower_ws = wb_values["10-Manpower Required"]
    manpower_rows = []
    for row_idx in range(4, min(manpower_ws.max_row, 40) + 1):
        if manpower_ws[f"B{row_idx}"].value is None:
            continue
        manpower_rows.append(
            {
                "awg": manpower_ws[f"B{row_idx}"].value,
                "secao_mm2": manpower_ws[f"C{row_idx}"].value,
                "descricao": _clean_text(manpower_ws[f"E{row_idx}"].value),
                "vpmo_lp": manpower_ws[f"L{row_idx}"].value,
                "uom": manpower_ws[f"N{row_idx}"].value,
                "si_metros_1": manpower_ws[f"P{row_idx}"].value,
                "si_metros_2": manpower_ws[f"Q{row_idx}"].value,
                "price_1": manpower_ws[f"R{row_idx}"].value,
                "price_2": manpower_ws[f"S{row_idx}"].value,
                "price_3": manpower_ws[f"T{row_idx}"].value,
            }
        )
    msi_rows = []
    for sheet_name in ("2019-MSI", "2020-MSI"):
        ws = wb_values[sheet_name]
        header_row = 2 if sheet_name == "2019-MSI" else 4
        headers = [_clean_text(v) for v in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
        for row in ws.iter_rows(min_row=header_row + 2, max_row=min(ws.max_row, header_row + 80), values_only=True):
            if not row[1]:
                continue
            data = {headers[idx] or f"col_{idx+1}": row[idx] for idx in range(min(len(headers), len(row)))}
            data["sheet"] = sheet_name
            msi_rows.append(data)
    return {
        "catalogo_cabos": pd.DataFrame(cable_rows),
        "catalogo_manpower_aux": pd.DataFrame(manpower_rows),
        "linhas_msi_legadas": pd.DataFrame(msi_rows),
    }


def extract_io_taxonomy(wb_values) -> pd.DataFrame:
    ws = wb_values["2020 - EKL LISTA IOs - PHC"]
    rows = []
    for row_idx in range(6, min(ws.max_row, 140) + 1):
        code = ws[f"F{row_idx}"].value
        io_code = _clean_text(ws[f"G{row_idx}"].value)
        desc = _clean_text(ws[f"H{row_idx}"].value)
        val = ws[f"I{row_idx}"].value
        if not code and not io_code and not desc:
            continue
        rows.append({"codigo": code, "io_code": io_code, "descricao": desc, "valor_ref": val, "row": row_idx})
    return pd.DataFrame(rows)


def write_report(inventory: list[dict[str, Any]]) -> None:
    lines = [
        "# Extração do workbook EKL-ORC_BASE_V1.6",
        "",
        "## O que foi extraído",
        "- parâmetros comerciais e operacionais do modelo central",
        "- tarifas de mão de obra do modelo e benchmarks de 2020",
        "- taxonomia de custos e secções de orçamento",
        "- linhas tipo do mapa cliente",
        "- resumos orçamentais monitorizados",
        "- catálogos auxiliares de cabos, manpower e linhas MSI",
        "- taxonomia de inputs/outputs do modelo",
        "",
        "## O que foi ignorado",
        "- folhas de rascunho (`g3-*`)",
        "- versões intermédias do modelo usadas apenas como histórico",
        "- folhas de cálculo auxiliar sem valor direto imediato no motor",
        "",
        "## Recomendações para integrar já",
        "- usar `parametros_comerciais.json` para defaults comerciais e logísticos",
        "- usar `tarifas_mao_obra.json` como referência de perfis e tarifário legado",
        "- usar `categorias_orcamento.json` e `linhas_tipo_orcamento.csv` para taxonomia e catálogos estruturados",
        "- usar `resumo_orcamental.json` apenas como mapa de saídas relevantes, não como motor de cálculo",
        "",
        "## Ficar só como referência",
        "- bases EN-US originais",
        "- folhas de conversão cambial e US->PT como suporte metodológico",
        "- folhas históricas 2016-2018 e versões intermédias do modelo",
        "",
        "## Mapeamento por folha",
    ]
    for row in inventory:
        lines.append(f"- `{row['sheet']}` -> `{row['decision']}`: {row['reason']}")
    REPORT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb_values = load_workbook(DEFAULT_WORKBOOK_PATH, data_only=True, read_only=True)
    wb_formulas = load_workbook(DEFAULT_WORKBOOK_PATH, data_only=False, read_only=True)

    inventory = workbook_inventory()
    _json_dump(OUTPUT_DIR / "ekl_workbook_inventory.json", inventory)
    _json_dump(OUTPUT_DIR / "parametros_comerciais.json", extract_parametros_comerciais(wb_values, wb_formulas))
    _json_dump(OUTPUT_DIR / "tarifas_mao_obra.json", extract_tarifas_mao_obra(wb_values))
    _json_dump(OUTPUT_DIR / "categorias_orcamento.json", extract_categorias_orcamento(wb_values))
    _json_dump(OUTPUT_DIR / "resumo_orcamental.json", extract_resumo_orcamental(wb_values))
    extract_linhas_tipo_orcamento(wb_values).to_csv(OUTPUT_DIR / "linhas_tipo_orcamento.csv", index=False)
    aux = extract_catalogos_auxiliares(wb_values)
    aux["catalogo_cabos"].to_csv(OUTPUT_DIR / "catalogo_cabos.csv", index=False)
    aux["catalogo_manpower_aux"].to_csv(OUTPUT_DIR / "catalogo_manpower_aux.csv", index=False)
    aux["linhas_msi_legadas"].to_csv(OUTPUT_DIR / "linhas_msi_legadas.csv", index=False)
    extract_io_taxonomy(wb_values).to_csv(OUTPUT_DIR / "taxonomia_ios.csv", index=False)
    write_report(inventory)
    print(OUTPUT_DIR)


if __name__ == "__main__":
    main()
