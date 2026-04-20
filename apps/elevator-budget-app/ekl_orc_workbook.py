from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook


DEFAULT_WORKBOOK_PATH = Path("/Users/edmundofrazao/Downloads/EKL-ORC_BASE_V1.6.xlsx")

MODEL_SHEET = "2020 - EKL ORC MODEL 0.5"
CLIENT_SHEET = "2020 - Ex ORC Cliente."
OVERVIEW_SHEET = "2020 - EKL COST OVERVIEW 0.3"


@dataclass(frozen=True)
class GuidedField:
    section: str
    sheet: str
    cell: str
    label: str
    input_type: str = "text"
    step: float | None = None
    help_text: str = ""

    @property
    def key(self) -> str:
        return f"{self.sheet}!{self.cell}"


GUIDED_FIELDS: tuple[GuidedField, ...] = (
    GuidedField("Projeto", CLIENT_SHEET, "A1", "Empresa"),
    GuidedField("Projeto", CLIENT_SHEET, "A2", "Cliente"),
    GuidedField("Projeto", CLIENT_SHEET, "A3", "Nome da empreitada"),
    GuidedField("Projeto", CLIENT_SHEET, "A4", "Título do mapa"),
    GuidedField("Projeto", CLIENT_SHEET, "A5", "Tipo de empreitada"),
    GuidedField("Projeto", CLIENT_SHEET, "A6", "Documento cliente"),
    GuidedField("Parâmetros do modelo", MODEL_SHEET, "O21", "Margem material", "number", 0.01),
    GuidedField("Parâmetros do modelo", MODEL_SHEET, "P21", "Margem mão de obra", "number", 0.01),
    GuidedField("Parâmetros do modelo", MODEL_SHEET, "Q21", "Desconto cliente", "number", 0.01),
    GuidedField("Parâmetros do modelo", MODEL_SHEET, "T21", "Número de homens", "number", 1.0),
    GuidedField("Parâmetros do modelo", MODEL_SHEET, "U21", "Horas semanais", "number", 1.0),
    GuidedField("Tarifas MO", MODEL_SHEET, "AV8", "Tarifa ajudante", "number", 0.5),
    GuidedField("Tarifas MO", MODEL_SHEET, "AV9", "Tarifa serralheiro instrumentos", "number", 0.5),
    GuidedField("Tarifas MO", MODEL_SHEET, "AV10", "Tarifa oficial electricista", "number", 0.5),
    GuidedField("Tarifas MO", MODEL_SHEET, "AV11", "Tarifa instrumentista", "number", 0.5),
    GuidedField("Tarifas MO", MODEL_SHEET, "AV12", "Tarifa encarregado", "number", 0.5),
    GuidedField("Tarifas MO", MODEL_SHEET, "AV13", "Tarifa técnico de segurança", "number", 0.5),
    GuidedField("Tarifas MO", MODEL_SHEET, "AV14", "Tarifa engenheiro", "number", 0.5),
    GuidedField("Tarifas MO", MODEL_SHEET, "AV15", "Tarifa gestor de projetos", "number", 0.5),
    GuidedField("Tarifas MO", MODEL_SHEET, "AV16", "Tarifa diretor de obra", "number", 0.5),
)

SUMMARY_CELLS: tuple[tuple[str, str, str], ...] = (
    (CLIENT_SHEET, "F12", "Total cliente"),
    (MODEL_SHEET, "Q22", "Total com desconto"),
    (MODEL_SHEET, "Q23", "Total sem desconto"),
    (MODEL_SHEET, "AS21", "Homem-hora total"),
    (OVERVIEW_SHEET, "M14", "Execução prevista (dias)"),
    (OVERVIEW_SHEET, "M26", "Custos totais MO"),
    (OVERVIEW_SHEET, "M32", "Vendas totais MO"),
    (OVERVIEW_SHEET, "M38", "Custos viaturas"),
    (OVERVIEW_SHEET, "M44", "Custos pernoita"),
    (OVERVIEW_SHEET, "M50", "Ajudas de custo"),
)


def workbook_exists() -> bool:
    return DEFAULT_WORKBOOK_PATH.exists()


def source_label(upload_name: str | None = None) -> str:
    return upload_name or DEFAULT_WORKBOOK_PATH.name


def load_workbook_bytes(upload_bytes: bytes | None = None) -> bytes:
    if upload_bytes is not None:
        return upload_bytes
    return DEFAULT_WORKBOOK_PATH.read_bytes()


def open_workbook(raw_bytes: bytes, *, data_only: bool) -> Workbook:
    return load_workbook(BytesIO(raw_bytes), data_only=data_only)


def read_cell(workbook: Workbook, sheet_name: str, cell: str) -> Any:
    return workbook[sheet_name][cell].value


def extract_guided_defaults(workbook: Workbook) -> dict[str, Any]:
    values: dict[str, Any] = {}
    for field in GUIDED_FIELDS:
        values[field.key] = read_cell(workbook, field.sheet, field.cell)
    return values


def extract_summary(workbook: Workbook) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for sheet_name, cell, label in SUMMARY_CELLS:
        rows.append(
            {
                "label": label,
                "sheet": sheet_name,
                "cell": cell,
                "value": read_cell(workbook, sheet_name, cell),
            }
        )
    return rows


def quantity_rows(workbook: Workbook, limit: int = 80) -> pd.DataFrame:
    ws = workbook[CLIENT_SHEET]
    rows: list[dict[str, Any]] = []
    for row_idx in range(1, min(ws.max_row, 180) + 1):
        qty = ws[f"D{row_idx}"].value
        if not isinstance(qty, (int, float)):
            continue
        rows.append(
            {
                "item": ws[f"A{row_idx}"].value or "",
                "descricao": ws[f"B{row_idx}"].value or "",
                "unidade": ws[f"C{row_idx}"].value or "",
                "quantidade": qty,
                "preco_unitario_ref": ws[f"E{row_idx}"].value or "",
                "total_formula": ws[f"F{row_idx}"].value or "",
                "cell": f"D{row_idx}",
            }
        )
    frame = pd.DataFrame(rows)
    if frame.empty:
        return frame
    return frame.head(limit)


def _coerce_scalar(value: Any, original: Any) -> Any:
    if value in ("", None):
        return None
    if isinstance(original, bool):
        text = str(value).strip().lower()
        return text in {"1", "true", "sim", "yes", "y"}
    if isinstance(original, int) and not isinstance(original, bool):
        return int(float(value))
    if isinstance(original, float):
        return float(value)
    return value


def apply_overrides(workbook: Workbook, overrides: dict[str, Any]) -> None:
    for key, raw_value in overrides.items():
        if "!" not in key:
            continue
        sheet_name, cell = key.split("!", 1)
        ws = workbook[sheet_name]
        original = ws[cell].value
        ws[cell] = _coerce_scalar(raw_value, original)


def workbook_to_bytes(workbook: Workbook) -> bytes:
    calc = workbook.calculation
    calc.calcMode = "auto"
    calc.fullCalcOnLoad = True
    calc.forceFullCalc = True
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_with_overrides(raw_bytes: bytes, overrides: dict[str, Any]) -> bytes:
    workbook = open_workbook(raw_bytes, data_only=False)
    apply_overrides(workbook, overrides)
    return workbook_to_bytes(workbook)


def preview_sheet(
    workbook: Workbook,
    sheet_name: str,
    *,
    max_rows: int = 120,
    max_cols: int = 18,
) -> pd.DataFrame:
    ws = workbook[sheet_name]
    row_count = min(ws.max_row, max_rows)
    col_count = min(ws.max_column, max_cols)
    data: list[list[Any]] = []
    for row_idx in range(1, row_count + 1):
        row_values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, col_count + 1)]
        data.append(row_values)
    columns = [get_column_letter(col_idx) for col_idx in range(1, col_count + 1)]
    frame = pd.DataFrame(data, columns=columns)
    frame.insert(0, "row", range(1, row_count + 1))
    return frame


def parse_override_rows(frame: pd.DataFrame) -> dict[str, Any]:
    overrides: dict[str, Any] = {}
    for row in frame.to_dict("records"):
        sheet_name = str(row.get("sheet", "")).strip()
        cell = str(row.get("cell", "")).strip().upper()
        if not sheet_name or not cell:
            continue
        overrides[f"{sheet_name}!{cell}"] = row.get("value")
    return overrides


def build_change_report(
    workbook: Workbook,
    overrides: dict[str, Any],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for key, new_value in overrides.items():
        if "!" not in key:
            continue
        sheet_name, cell = key.split("!", 1)
        current_value = read_cell(workbook, sheet_name, cell)
        rows.append(
            {
                "sheet": sheet_name,
                "cell": cell,
                "current_value": current_value,
                "new_value": new_value,
            }
        )
    frame = pd.DataFrame(rows)
    if frame.empty:
        return frame
    return frame.sort_values(["sheet", "cell"]).reset_index(drop=True)
