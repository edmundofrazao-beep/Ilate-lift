from __future__ import annotations

import io
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADER_FILL = PatternFill("solid", fgColor="16324F")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _format_currency_columns(ws, headers: list[str]) -> None:
    currency_tokens = ("custo", "mao_obra", "materiais", "transport", "auxiliares", "directo", "subtotal", "iva", "total")
    for idx, header in enumerate(headers, start=1):
        if any(token in header for token in currency_tokens):
            for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for item in cell:
                    item.number_format = '#,##0.00 [$EUR]'


def _format_hours_columns(ws, headers: list[str]) -> None:
    for idx, header in enumerate(headers, start=1):
        if "horas" in header:
            for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for item in cell:
                    item.number_format = "0.00"


def _add_table(ws, name: str, end_col: int, end_row: int) -> None:
    if end_row < 2 or end_col < 1:
        return
    table = Table(displayName=name, ref=f"A1:{get_column_letter(end_col)}{end_row}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)


def _write_dataframe(ws, frame: pd.DataFrame, title: str) -> None:
    if frame.empty:
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=12)
        ws["A3"] = "Sem dados."
        return
    headers = [str(col) for col in frame.columns]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for row in frame.itertuples(index=False):
        ws.append(list(row))
    _format_currency_columns(ws, headers)
    _format_hours_columns(ws, headers)
    _add_table(ws, f"Tabela{ws.title}".replace(" ", ""), len(headers), len(frame) + 1)
    ws.freeze_panes = "A2"
    for col in ws.columns:
        length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 12), 38)


def build_export_workbook(
    state: dict[str, Any],
    results: dict[str, Any],
    alerts: list[dict[str, str]],
) -> bytes:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumo"

    ws_summary["A1"] = "Orçamentação Técnica de Elevadores"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A3"] = "Cliente"
    ws_summary["B3"] = state["project"]["cliente"]
    ws_summary["A4"] = "Obra / Equipamento"
    ws_summary["B4"] = state["project"]["obra"]
    ws_summary["A5"] = "N.º proposta"
    ws_summary["B5"] = state["project"]["numero_proposta"]
    ws_summary["A6"] = "Tipo de orçamento"
    ws_summary["B6"] = state["project"]["tipo_orcamento"]
    ws_summary["A7"] = "Data exportação"
    ws_summary["B7"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    total_rows = [
        ("Horas totais", results["totals"]["horas"]),
        ("Mão de obra", results["totals"]["mao_obra"]),
        ("Materiais", results["totals"]["materiais"]),
        ("Custos invisíveis", results["totals"]["invisiveis"]),
        ("Custo directo", results["totals"]["directo"]),
        ("Subtotal s/ IVA", results["totals"]["subtotal_sem_iva"]),
        ("IVA", results["totals"]["iva"]),
        ("Preço final", results["totals"]["total_final"]),
    ]
    start_row = 10
    for idx, (label, value) in enumerate(total_rows, start=start_row):
        ws_summary[f"A{idx}"] = label
        ws_summary[f"B{idx}"] = value
        ws_summary[f"B{idx}"].number_format = '#,##0.00 [$EUR]' if idx != start_row else "0.00"
    ws_summary["B9"].number_format = "0.00"
    ws_summary["D3"] = "Alertas"
    ws_summary["D3"].font = Font(bold=True)
    if alerts:
        for idx, alert in enumerate(alerts, start=4):
            ws_summary[f"D{idx}"] = f"[{alert['level'].upper()}] {alert['text']}"
    else:
        ws_summary["D4"] = "Sem alertas."
    ws_summary.column_dimensions["A"].width = 24
    ws_summary.column_dimensions["B"].width = 16
    ws_summary.column_dimensions["D"].width = 80

    project_df = pd.DataFrame(
        [{"campo": key, "valor": value} for key, value in state["project"].items()]
        + [{"campo": f"tecnico_{key}", "valor": value} for key, value in state["technical"].items()]
        + [{"campo": f"custos_{key}", "valor": value} for key, value in state["costs"].items()]
    )
    ws_project = wb.create_sheet("Dados_Projeto")
    _write_dataframe(ws_project, project_df, "Dados do projeto")

    ws_ops = wb.create_sheet("Operacoes")
    _write_dataframe(ws_ops, results["operations"], "Operações orçamentadas")

    ws_extras = wb.create_sheet("Extras")
    _write_dataframe(ws_extras, results["extras"], "Custos adicionais")

    snapshot_df = pd.DataFrame(
        [{"codigo": code, **values} for code, values in state["operation_inputs"].items() if code in state["selected_codes"]]
    )
    ws_cfg = wb.create_sheet("Configuracao_Operacoes")
    _write_dataframe(ws_cfg, snapshot_df, "Configuração das operações")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
