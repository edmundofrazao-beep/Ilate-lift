from __future__ import annotations

from typing import Any

import pandas as pd
from openpyxl import load_workbook

from elevator_budget_app.config import SOURCE_WORKBOOK


def _normalise_cell(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def _sheet_to_frame(workbook, sheet_name: str, header_row: int) -> pd.DataFrame:
    ws = workbook[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [_normalise_cell(v) for v in rows[header_row - 1]]
    frame = pd.DataFrame(rows[header_row:], columns=headers)
    first_col = frame.columns[0]
    frame = frame[frame[first_col].notna()].copy()
    frame.columns = [str(col).strip() for col in frame.columns]
    for column in frame.columns:
        frame[column] = frame[column].map(_normalise_cell)
    return frame.reset_index(drop=True)


def load_catalogs() -> dict[str, Any]:
    workbook = load_workbook(SOURCE_WORKBOOK, data_only=True, read_only=True)

    parametros_df = _sheet_to_frame(workbook, "Parametros_PT", 1)
    tipos_df = _sheet_to_frame(workbook, "Tipos_Intervencao", 1)
    fatores_df = _sheet_to_frame(workbook, "Fatores_PT", 1)
    perfis_df = _sheet_to_frame(workbook, "Perfis_Equipa", 1)
    equipas_df = _sheet_to_frame(workbook, "Equipas_PT", 1)
    operacoes_df = _sheet_to_frame(workbook, "Operacoes_Detalhadas", 3)
    manutencao_df = _sheet_to_frame(workbook, "Operacoes_Manutencao", 4)
    micro_df = _sheet_to_frame(workbook, "Biblioteca_MicroMateriais", 3)
    logistica_df = _sheet_to_frame(workbook, "Biblioteca_Logistica_Viaturas", 3)
    fabricantes_df = _sheet_to_frame(workbook, "Fabricantes_Gamas_v15", 3)
    kits_operacoes_df = _sheet_to_frame(workbook, "Kits_Operacoes", 3)
    kits_fabricante_df = _sheet_to_frame(workbook, "Kits_Fabricante_v15", 3)
    arquitetura_df = _sheet_to_frame(workbook, "Arquitetura_Elevador", 3)
    subsistemas_df = _sheet_to_frame(workbook, "Subsistemas_Elevador", 3)
    componentes_df = _sheet_to_frame(workbook, "Componentes_Elevador", 3)

    parametros = {
        str(row["Parâmetro"]): float(row["Valor"])
        for _, row in parametros_df.iterrows()
        if pd.notna(row["Parâmetro"]) and pd.notna(row["Valor"])
    }
    tipo_fatores = {
        str(row["Tipo de intervenção"]): float(row["Fator por defeito"])
        for _, row in tipos_df.iterrows()
        if pd.notna(row["Tipo de intervenção"]) and pd.notna(row["Fator por defeito"])
    }

    dificuldade_fatores = {
        str(row["Nível"]): float(row["Fator"])
        for _, row in fatores_df.iterrows()
        if row["Grupo"] == "Dificuldade técnica"
    }
    logistica_fatores = {
        str(row["Nível"]): float(row["Fator"])
        for _, row in fatores_df.iterrows()
        if row["Grupo"] == "Logística e acesso"
    }

    taxas_perfil = {
        str(row["Perfil"]): float(row["Taxa horária (€ / h)"])
        for _, row in perfis_df.iterrows()
        if pd.notna(row["Perfil"]) and pd.notna(row["Taxa horária (€ / h)"])
    }

    equipas_df = equipas_df.copy()
    equipas_df["Custo hora calculado"] = (
        equipas_df["N.º Oficiais"].fillna(0) * taxas_perfil.get("Oficial", 0.0)
        + equipas_df["N.º Ajudantes"].fillna(0) * taxas_perfil.get("Ajudante", 0.0)
        + equipas_df["N.º Especialistas"].fillna(0) * taxas_perfil.get("Especialista", 0.0)
        + equipas_df["N.º Engenheiros"].fillna(0) * taxas_perfil.get("Engenheiro", 0.0)
    )
    taxas_equipa = {
        str(row["Equipa"]): float(row["Custo hora calculado"])
        for _, row in equipas_df.iterrows()
        if pd.notna(row["Equipa"])
    }

    operacoes_df = operacoes_df.rename(
        columns={
            "Código": "codigo",
            "Fase": "fase",
            "Operação": "operacao",
            "Unidade": "unidade",
            "Quantidade base": "quantidade_base",
            "Tempo fixo": "tempo_fixo",
            "Tempo variável": "tempo_variavel",
            "Equipa": "equipa",
            "Material unitário": "material_unitario",
            "Consumíveis fixos": "consumiveis_fixos",
            "Transporte": "transporte_base",
            "Equipamento auxiliar": "equipamento_auxiliar_base",
            "Subcontratos": "subcontrato_base",
            "Observações": "observacoes",
            "Tipo de intervenção": "tipo_intervencao",
            "Sistema técnico": "sistema",
            "Sub-sistema técnico": "sub_sistema",
            "Componente principal": "componente",
            "Kit micro-materiais sugerido": "kit_micro",
            "Sugestão micro (€)": "micro_sugerido",
            "Kit logística sugerido": "kit_logistica",
            "Sugestão logística (€)": "logistica_sugerida",
            "Sugestão meios auxiliares (€)": "meios_auxiliares_sugeridos",
            "Sugestão total invisível (€)": "invisiveis_sugeridos",
        }
    )
    keep_cols = [
        "codigo",
        "fase",
        "operacao",
        "unidade",
        "quantidade_base",
        "tempo_fixo",
        "tempo_variavel",
        "equipa",
        "material_unitario",
        "consumiveis_fixos",
        "transporte_base",
        "equipamento_auxiliar_base",
        "subcontrato_base",
        "observacoes",
        "tipo_intervencao",
        "sistema",
        "sub_sistema",
        "componente",
        "kit_micro",
        "micro_sugerido",
        "kit_logistica",
        "logistica_sugerida",
        "meios_auxiliares_sugeridos",
        "invisiveis_sugeridos",
    ]
    operacoes_df = operacoes_df[keep_cols].copy()
    operacoes_df = operacoes_df[operacoes_df["codigo"].notna()].reset_index(drop=True)

    manutencao_df = manutencao_df.rename(
        columns={
            "Código": "codigo",
            "Fase": "fase",
            "Operação": "operacao",
            "Unidade": "unidade",
            "Tempo fixo (h)": "tempo_fixo",
            "Equipa": "equipa",
            "Consumíveis (€)": "consumiveis_fixos",
            "Sub-sistema": "sub_sistema",
        }
    )
    if not manutencao_df.empty:
        manutencao_df["quantidade_base"] = 1.0
        manutencao_df["tempo_variavel"] = 0.0
        manutencao_df["material_unitario"] = 0.0
        manutencao_df["transporte_base"] = 0.0
        manutencao_df["equipamento_auxiliar_base"] = 0.0
        manutencao_df["subcontrato_base"] = 0.0
        manutencao_df["tipo_intervencao"] = "Manutenção preventiva"
        manutencao_df["sistema"] = "Manutenção"
        manutencao_df["componente"] = manutencao_df["sub_sistema"]
        manutencao_df["kit_micro"] = ""
        manutencao_df["micro_sugerido"] = manutencao_df["consumiveis_fixos"].fillna(0.0)
        manutencao_df["kit_logistica"] = ""
        manutencao_df["logistica_sugerida"] = 0.0
        manutencao_df["meios_auxiliares_sugeridos"] = 0.0
        manutencao_df["invisiveis_sugeridos"] = manutencao_df["micro_sugerido"].fillna(0.0)
        manutencao_df["observacoes"] = "Catálogo de manutenção preventiva"
        operacoes_df = pd.concat([operacoes_df, manutencao_df[keep_cols]], ignore_index=True)

    operacoes_df = operacoes_df.drop_duplicates(subset=["codigo"], keep="first").reset_index(drop=True)
    numeric_cols = [
        "quantidade_base",
        "tempo_fixo",
        "tempo_variavel",
        "material_unitario",
        "consumiveis_fixos",
        "transporte_base",
        "equipamento_auxiliar_base",
        "subcontrato_base",
        "micro_sugerido",
        "logistica_sugerida",
        "meios_auxiliares_sugeridos",
        "invisiveis_sugeridos",
    ]
    for col in numeric_cols:
        operacoes_df[col] = pd.to_numeric(operacoes_df[col], errors="coerce").fillna(0.0)

    return {
        "parametros": parametros,
        "tipo_fatores": tipo_fatores,
        "dificuldade_fatores": dificuldade_fatores,
        "logistica_fatores": logistica_fatores,
        "taxas_perfil": taxas_perfil,
        "taxas_equipa": taxas_equipa,
        "equipas_df": equipas_df,
        "operacoes_df": operacoes_df,
        "micro_df": micro_df,
        "logistica_df": logistica_df,
        "fabricantes_df": fabricantes_df,
        "kits_operacoes_df": kits_operacoes_df,
        "kits_fabricante_df": kits_fabricante_df,
        "arquitetura_df": arquitetura_df,
        "subsistemas_df": subsistemas_df,
        "componentes_df": componentes_df,
    }
